# speech recognition

import speech_recognition as speech
from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active


#a lot of variables used for loops and conditions
voice = speech.Recognizer()

while True:
    #with microphone input as the input
    with speech.Microphone() as source:

        print("Say something!")

        #variable = the input from the microphone
        audio = voice.listen(source)
        voiceRecog = voice.recognize_google(audio)
        try:
            #print the word that was heard
            print("\nYou said: " + voiceRecog)
        except speech.UnknownValueError:
            #if couldnt recognise then print this msg
            print("Sorry, we couldn't make that out. Try again!")

        except speech.RequestError as e:
            print("Could not request results from Google Speech Recognition service")


        #list of commands  for comparing   
        words = voiceRecog.split(" ")
        print(words)
        if "apply" in voiceRecog:
            break
        elif words[0] == "add":
            i = 1
            p1 = words[1]
            p2 = words[2]
            p3 = words[4]
            print(p1 + p2 + p3)
            print("\nCommand Found!\n")


        elif words[0] == "absolute":
            i=2
            p1 = words[1]
            p2 = words[3]
            print("\nCommand Found!\n")
        else:
            print("\nCommand not found!\n\n")

print("\n Done")

if i==1:
    ws[p3] = '=SUM('+p1+':'+p2+')'
elif i==2:
    ws[p2] = '=ABS('+p1+')'

# Save the file
wb.save("sample.xlsx")
